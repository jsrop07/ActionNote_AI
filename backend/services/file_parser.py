# backend/services/file_parser.py
# 역할:
# - 업로드된 파일을 확장자별로 읽어 텍스트로 변환합니다.
# - 파일 입력도 최종적으로 analyze_text(text) 함수에 전달할 수 있도록 문자열로 정규화합니다.
# - 1차 구현 순서는 TXT → DOCX → PDF → XLSX입니다.
import fitz 
import openpyxl
from io import BytesIO
from docx import Document
from fastapi import UploadFile


# 이 함수는 업로드된 파일명에서 확장자를 추출합니다.
# 확장자에 따라 TXT, DOCX, PDF, XLSX 파서를 선택하기 위해 사용합니다.
def get_file_extension(filename: str) -> str:
    return filename.lower().split(".")[-1]


# 이 함수는 TXT 파일에서 텍스트를 추출합니다.
# UTF-8을 우선 시도하고, 실패하면 CP949로 다시 시도해 한글 윈도우 텍스트 파일도 처리합니다.
async def parse_txt_file(file: UploadFile) -> str:
    file_bytes = await file.read()

    try:
        return file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return file_bytes.decode("cp949", errors="ignore")


# 이 함수는 DOCX 파일에서 문단 텍스트를 추출합니다.
# 일반적인 회의록/업무 메모 문서의 paragraph 내용을 문자열로 합칩니다.
async def parse_docx_file(file: UploadFile) -> str:
    file_bytes = await file.read()
    document = Document(BytesIO(file_bytes))

    paragraphs = []

    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if text:
            paragraphs.append(text)

    return "\n".join(paragraphs)


# 이 함수는 PDF 파일에서 페이지별 텍스트를 추출합니다.
# 텍스트 기반 PDF는 처리 가능하지만, 스캔 이미지 PDF는 OCR이 필요하므로 이번 범위에서는 제외합니다.
async def parse_pdf_file(file: UploadFile) -> str:
    file_bytes = await file.read()

    pdf = fitz.open(stream=file_bytes, filetype="pdf")
    pages = []

    for page_index, page in enumerate(pdf, start=1):
        page_text = page.get_text().strip()

        if page_text:
            pages.append(f"[Page {page_index}]\n{page_text}")

    return "\n\n".join(pages)


# 이 함수는 XLSX 파일에서 모든 시트의 셀 값을 텍스트로 추출합니다.
# 표 형태 회의록도 AI가 분석할 수 있도록 시트명과 행 단위 텍스트로 변환합니다.
async def parse_xlsx_file(file: UploadFile) -> str:
    file_bytes = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    lines = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines.append(f"[Sheet: {sheet_name}]")

        for row in sheet.iter_rows(values_only=True):
            values = [
                str(cell).strip()
                for cell in row
                if cell is not None and str(cell).strip()
            ]

            if values:
                lines.append(" | ".join(values))

        lines.append("")

    return "\n".join(lines).strip()


# 이 함수는 업로드된 파일의 확장자를 보고 알맞은 파서를 실행합니다.
# 현재 1차로 TXT를 먼저 연결하고, 이후 DOCX/PDF/XLSX를 순서대로 추가합니다.
async def parse_uploaded_file(file: UploadFile) -> str:
    extension = get_file_extension(file.filename or "")

    if extension == "txt":
        text = await parse_txt_file(file)
    elif extension == "docx":
        text = await parse_docx_file(file)
    elif extension == "pdf":
        text = await parse_pdf_file(file)
    elif extension == "xlsx":
        text = await parse_xlsx_file(file)
    else:
        raise ValueError(f"지원하지 않는 파일 형식입니다: .{extension}")

    if not text or not text.strip():
        raise ValueError("파일에서 텍스트를 추출하지 못했습니다.")

    return text